from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame

Data = List[List[Union[str, Decimal]]]


def configure_worksheet_columns(
    worksheet: Worksheet,
    column_widths: Optional[Dict[str, int]] = None,
    style_columns: Optional[List[str]] = None,
    bold_columns: Optional[List[str]] = None,
    merge_columns: Optional[List[str]] = None,
) -> None:
    border_style: Border = Border(
        Side(border_style="thin"),
        Side(border_style="thin"),
        Side(border_style="thin"),
        Side(border_style="thin"),
    )

    if column_widths:
        for column in worksheet.columns:
            column_letter: str = column[0].column_letter
            worksheet.column_dimensions[column_letter].width = column_widths[column_letter]

            if style_columns:
                for cell in column:
                    if cell.row > 1 and column_letter in style_columns:
                        cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border_style

    last_row: int = worksheet.max_row

    if bold_columns:
        for cell in worksheet[last_row]:
            if cell.column_letter in bold_columns:
                cell.font = Font(bold=True)

    if merge_columns:
        worksheet.merge_cells(f"{merge_columns[0]}{last_row}:{merge_columns[-1]}{last_row}")


def write_data_to_worksheet(
    workbook: Workbook,
    sheet_name: str,
    headers: List[str],
    data: Data,
    column_widths: Optional[Dict[str, int]] = None,
    style_columns: Optional[List[str]] = None,
    filter_columns: Optional[List[str]] = None,
    bold_columns: Optional[List[str]] = None,
    merge_columns: Optional[List[str]] = None,
) -> None:
    dataframe: DataFrame = DataFrame(data=data, columns=headers)
    worksheet: Worksheet = workbook.create_sheet()

    if sheet_name:
        worksheet.title = sheet_name

    for row in dataframe_to_rows(dataframe, index=False, header=True):
        worksheet.append(row)

    filter_range: str = worksheet.dimensions

    if filter_columns:
        filter_range: str = f"{filter_columns[0]}1:{filter_columns[-1]}1"

    worksheet.auto_filter.ref = filter_range

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    configure_worksheet_columns(
        worksheet=worksheet,
        column_widths=column_widths,
        style_columns=style_columns,
        bold_columns=bold_columns,
        merge_columns=merge_columns,
    )


def get_report_file(
    tasks_data: Data = [],
    employees_data: Data = [],
    basic_orders_data: Data = [],
    detailed_orders_data: Data = [],
) -> BytesIO:
    workbook: Workbook = Workbook()

    workbook.remove(workbook.active)

    write_data_to_worksheet(
        workbook=workbook,
        sheet_name="Назначенные задания",
        headers=[
            "ФИО сотрудника",
            "Таб. номер",
            "Категория сотрудника",
            "Наименование подразделения",
            "Номер заказа",
            "Наименование заказа",
            "Наименование работы",
            "Дата выполнения",
            "Затраченное время, ч",
        ],
        data=tasks_data,
        column_widths={"A": 28, "B": 18, "C": 18, "D": 22, "E": 22, "F": 44, "G": 44, "H": 24, "I": 18},
        style_columns=["I"],
        filter_columns=["A", "B", "C", "D", "E", "F", "G", "H"],
    )

    write_data_to_worksheet(
        workbook=workbook,
        sheet_name="Табель рабочего времени",
        headers=[
            "ФИО сотрудника",
            "Таб. номер",
            "Категория сотрудника",
            "Наименование подразделения",
            "Дата выполнения",
            "Затраченное время, ч",
        ],
        data=employees_data,
        column_widths={"A": 28, "B": 18, "C": 18, "D": 22, "E": 24, "F": 18},
        style_columns=["F"],
        filter_columns=["A", "B", "C", "D", "E"],
    )

    write_data_to_worksheet(
        workbook=workbook,
        sheet_name="Сводка по заказам",
        headers=[
            "Номер заказа",
            "Наименование заказа",
            "Плановая трудоемкость, ч",
            "Фактическая трудоемкость, ч",
            "Остаточная трудоемкость, ч",
        ],
        data=basic_orders_data,
        column_widths={"A": 22, "B": 44, "C": 22, "D": 22, "E": 22},
        style_columns=["C", "D", "E"],
        filter_columns=["A", "B"],
        bold_columns=["A", "B", "C", "D", "E"],
        merge_columns=["A", "B"],
    )

    write_data_to_worksheet(
        workbook=workbook,
        sheet_name="Детализация по заказам",
        headers=[
            "Номер заказа",
            "Наименование заказа",
            "Наименование работы",
            "Плановая трудоемкость, ч",
            "Фактическая трудоемкость, ч",
            "Остаточная трудоемкость, ч",
        ],
        data=detailed_orders_data,
        column_widths={"A": 22, "B": 44, "C": 44, "D": 22, "E": 22, "F": 22},
        style_columns=["C", "D", "E", "F"],
        filter_columns=["A", "B", "C"],
    )

    file: BytesIO = BytesIO()
    workbook.save(file)
    file.seek(0)
    return file
